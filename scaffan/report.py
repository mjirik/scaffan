# /usr/bin/env python
# -*- coding: utf-8 -*-
"""
Process lobulus analysis.
"""
import logging

logger = logging.getLogger(__name__)
import pandas as pd
import os.path as op
import os
import matplotlib.pyplot as plt
import skimage.io
import warnings
from pathlib import Path
import numpy as np


class Report:
    def __init__(self, outputdir=None, additional_spreadsheet_fn=None):
        # self.outputdir = op.expanduser(outputdir)


        self.df: pd.DataFrame = None
        self.imgs = {}
        self.actual_row = {}
        self.show = False
        self.save = False
        self.debug = False
        self.level = 50
        self.additional_spreadsheet_fn = additional_spreadsheet_fn

        if outputdir is not None:
            self.set_output_dir(outputdir)

    def set_output_dir(self, outputdir):
        self.outputdir = Path(outputdir).expanduser()
        if not op.exists(self.outputdir):
            os.makedirs(self.outputdir)

        self.df = pd.DataFrame()
        self.imgs = {}
        self.actual_row = {}

    def set_show(self, show):
        self.show = show

    def set_save(self, save):
        self.save = save

    def add_cols_to_actual_row(self, data):
        self.actual_row.update(data)

    # def write_table(self, filename):
    def finish_actual_row(self):
        data = self.actual_row
        df = pd.DataFrame([list(data.values())], columns=list(data.keys()))
        self.df = self.df.append(df, ignore_index=True)

            # if excel_path.exists():
            #     print("append to excel")
            #
            # else:
            #
            #     # writer = pd.ExcelWriter(filename, engine='openpyxl')
            #     df.to_excel(filename)
            #     print("create new excel")

        self.actual_row = {}

    def add_table(self):
        pass

    def dump(self):
        self.df.to_excel(op.join(self.outputdir, "data.xlsx"))

        if self.additional_spreadsheet_fn is not None:
            excel_path = Path(self.additional_spreadsheet_fn)
            # print("we will write to excel", excel_path)
            filename = str(excel_path)
            append_df_to_excel(filename, self.df)
            # append_df_to_excel_no_head_processing(filename, self.df)
        self.df = pd.DataFrame()
        self.imgs = {}
        self.actual_row = {}

    def imsave(self, base_fn, arr, k=50):
        """
        :param base_fn: with a format slot for annotation id like "skeleton_{}.png"
        :param arr:
        :return:
        """

        if self.save:
            plt.imsave(op.join(self.outputdir, base_fn), arr)
        filename, ext = op.splitext(base_fn)
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", ".*low contrast image.*")
            # warnings.simplefilter("low contrast image")
            if self.save:
                skimage.io.imsave(op.join(self.outputdir, filename + "_raw" + ext), k * arr)
        self.imgs[base_fn] = arr

    def imsave_as_fig(self, base_fn, arr):
        filename, ext = op.splitext(base_fn)
        fig = plt.figure()
        plt.imshow(arr)
        plt.colorbar()
        if self.save:
            plt.savefig(op.join(self.outputdir, filename + "" + ext))
        if self.show:
            plt.show()
        else:
            plt.close(fig)

    # def add_array(self, base_fn, arr, k=50):
    #     if self.save:
    #         self.imsave

    def savefig_and_show(self, base_fn, fig):
        filename, ext = op.splitext(base_fn)
        if self.save:
            plt.savefig(op.join(self.outputdir, filename + "" + ext))
        if self.show:
            plt.show()
        else:
            plt.close(fig)

    def save_np_data(self, base_fn, data, format_args=None, level=60):
        if format_args is None:
            format_args = []
        fn = op.join(self.outputdir, base_fn.format(format_args))
        if self.level < level:
            np.save(data, fn)

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Excel file have to be closed.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd
    filename = Path(filename)
    if filename.exists():
        # writer = pd.ExcelWriter(filename, engine='openpyxl')


        dfold = pd.read_excel(str(filename), sheet_name=sheet_name)
        # dfout = pd.concat([dfin, df], axis=0, ignore_index=True)
        dfcombine = dfold.append(df, ignore_index=True, sort=True)
        dfcombine.to_excel(str(filename), sheet_name=sheet_name, index=False)
        # try:
        #     dfold = pd.read_excel(str(filename), sheet_name=sheet_name)
        #     dfcombine = dfold.append(df, ignore_index=True)
        #     dfcombine.to_excel(str(filename), sheet_name=sheet_name)
        # except PermissionError as e:
            # print("File is opened in other application")
            # import xlwings as xw
            # sht = xw.Book(str(filename)).sheets[0]
            # sht.range('A1').expand().options(pd.DataFrame).value


    else:
        # pd.read_excel(filename, sheet_name=)
        df.to_excel(str(filename), sheet_name=sheet_name, index=False)
        pass

    # # ignore [engine] parameter if it was passed
    # if 'engine' in to_excel_kwargs:
    #     to_excel_kwargs.pop('engine')
    #
    # writer = pd.ExcelWriter(filename, engine='openpyxl')
    #
    # # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    # try:
    #     FileNotFoundError
    # except NameError:
    #     FileNotFoundError = IOError
    #
    #
    # try:
    #     # try to open an existing workbook
    #     writer.book = load_workbook(filename)
    #
    #     # get the last row in the existing Excel sheet
    #     # if it was not specified explicitly
    #     if startrow is None and sheet_name in writer.book.sheetnames:
    #         startrow = writer.book[sheet_name].max_row
    #
    #     # truncate sheet
    #     if truncate_sheet and sheet_name in writer.book.sheetnames:
    #         # index of [sheet_name] sheet
    #         idx = writer.book.sheetnames.index(sheet_name)
    #         # remove [sheet_name]
    #         writer.book.remove(writer.book.worksheets[idx])
    #         # create an empty sheet [sheet_name] using old index
    #         writer.book.create_sheet(sheet_name, idx)
    #
    #     # copy existing sheets
    #     writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    # except FileNotFoundError:
    #     # file does not exist yet, we will create it
    #     pass
    #
    # if startrow is None:
    #     startrow = 0
    #
    # # write out the new sheet
    # df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    #
    # # save the workbook
    # writer.save()


def append_df_to_excel_no_head_processing(filename, df, sheet_name='Sheet1', startrow=None,
                                          truncate_sheet=False,
                                          **to_excel_kwargs):
    """

    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.
    It does insert also first line with column name :-(

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()